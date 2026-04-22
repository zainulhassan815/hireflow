"""Spreadsheet → table provider (F105.c).

Handles xlsx via openpyxl. Legacy .xls (needs xlrd) and .ods (needs
odfpy) are deliberately out of scope — add only when a real user
asks.

``prepare`` parses every sheet and writes JSON to
``viewable/<doc_id>.json``. ``render`` fetches that JSON back and
inlines it as ``data`` — the browser gets the table payload in the
same response that tells it how to render.
"""

from __future__ import annotations

import json
import logging
from io import BytesIO

from openpyxl import load_workbook

from app.adapters.protocols import BlobStorage
from app.adapters.viewers._table_shape import (
    TableSheet,
    store_table_payload,
    table_viewable_key,
    truncate,
)
from app.adapters.viewers.protocol import (
    PreparationResult,
    StorageGetSync,
    StoragePutSync,
    ViewablePayload,
)
from app.models import Document

logger = logging.getLogger(__name__)

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class SpreadsheetProvider:
    def accepts(self, mime_type: str | None) -> bool:
        return mime_type == _XLSX_MIME

    def prepare(
        self,
        doc: Document,
        *,
        storage_get: StorageGetSync,
        storage_put: StoragePutSync,
    ) -> PreparationResult:
        source = storage_get(doc.storage_key)
        sheets = self._parse(source)
        key = store_table_payload(doc_id=doc.id, sheets=sheets, storage_put=storage_put)
        logger.info(
            "spreadsheet → table: source_bytes=%d sheets=%d",
            len(source),
            len(sheets),
        )
        return PreparationResult(kind="table", key=key)

    async def render(self, doc: Document, storage: BlobStorage) -> ViewablePayload:
        if not doc.viewable_key:
            return ViewablePayload(
                kind="unsupported",
                meta={
                    "mime_type": doc.mime_type,
                    "filename": doc.filename,
                    "reason": "conversion_pending",
                },
            )
        blob = await storage.get(doc.viewable_key)
        data = json.loads(blob.decode("utf-8"))
        return ViewablePayload(
            kind="table",
            data=data,
            meta={
                "filename": doc.filename,
                "sheet_count": len(data.get("sheets", [])),
            },
        )

    @staticmethod
    def _parse(source_bytes: bytes) -> list[TableSheet]:
        # ``read_only=True`` streams rows without loading the whole
        # workbook into memory. ``data_only=True`` returns cached
        # formula values rather than formula strings — what a viewer
        # should show.
        workbook = load_workbook(BytesIO(source_bytes), read_only=True, data_only=True)
        sheets: list[TableSheet] = []
        try:
            for worksheet in workbook.worksheets:
                rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
                headers, data_rows, total_rows, total_cols = truncate(rows)
                sheets.append(
                    TableSheet(
                        name=worksheet.title,
                        headers=headers,
                        rows=data_rows,
                        total_rows=total_rows,
                        total_cols=total_cols,
                    )
                )
        finally:
            workbook.close()
        return sheets


__all__ = ["SpreadsheetProvider", "table_viewable_key"]
